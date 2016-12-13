
import openpyxl, zipfile, subprocess, time




wb1 = openpyxl.load_workbook('test.xlsx')

sheet1 = wb1.get_sheet_by_name('Sheet1')

stocksString = ''
#print(cell_range[2])

for x in range (2,21):
    #print(sheet1.cell(row=1,column=x).value)
    if (x != 20):
        stocksString = stocksString + sheet1.cell(row=1,column=x).value +'.AX+' 
    else: 
        stocksString = stocksString + sheet1.cell(row=1,column=x).value +'.AX' 

print (stocksString)