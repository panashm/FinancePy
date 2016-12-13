import openpyxl, zipfile, subprocess

import pandas as pd

from yahoo_finance import Share

ANZ = Share('ANZ.AX')

yahoo = Share('YHOO')

print (yahoo.get_price())

print ("printing the price now:");

ANZ.refresh()

print ANZ.get_price()

url = 'http://finance.yahoo.com/d/quotes.csv?s=AAPL+GOOG+MSFT&f=nab'


response = urllib2.urlopen(url)
cr = csv.reader(response)

for row in cr:
    print row


#   subprocess.Popen([r'/Applications/Microsoft Excel.app',  r'/Users/Panashe/excelPy/document.xlsx', shell=True])

#wb = openpyxl.load_workbook('invest.xltx')
#file = 'invest.xlsx'
#print (wb)
#sheet = wb.get_sheet_by_name('Sheet1')

#c = sheet['A1']

#print (c.value)

print ("hello world")

#old_sheet = wb.get_sheet_by_name('Sheet1')
#old_sheet.title = 'Sheet1.5'
#max_row = old_sheet.max_row
#max_col = old_sheet.max_column
#wb.create_sheet('Sheet1new', 0)

#print ("the max column is:", max_col)

#print ("the max row is:", max_row)

#new_sheet = wb.get_sheet_by_name('Sheet1new')

# Do the header.
#for col_num in range(1, max_col+1):
    #new_sheet.cell(row=1, column=col_num).value = old_sheet.cell(row=1, column=col_num).value

# The row to be inserted. We're manually populating each cell.

# Now do the rest of it. Note the row offset.
#for row_num in range(1, max_row+1):
   #for col_num in range (1, max_col+1):
        #new_sheet.cell(row = (row_num + 1), column = col_num).value = old_sheet.cell(row = row_num, column = col_num).value

#new_sheet.cell(row=1, column=1).value = 'DUMyyyMY'
#new_sheet.cell(row=1, column=2).value = 'DUMMyyyyY'
#wb.remove_sheet(old_sheet)
#new_sheet.title = 'Sheet1'


#wb.save(file)

#wb.save('invest.xlsx')

#wb.template = False
#wb.save('document.xlsx', as_template=False)

#with zipfile.ZipFile('/Users/Panashe/excelPy/document.xlsx', 'r') as zgood:
 #   styles_xml = zgood.read('xl/styles.xml')

#with zipfile.ZipFile('/Users/Panashe/excelPy/document.xlsx', 'a') as zbad:
#    zbad.writestr('xl/styles.xml', styles_xml)