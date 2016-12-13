import urllib3, pandas as pd

import openpyxl, zipfile, subprocess, time

#grab the stock headings from the worksheet, cells A1 to T1
#condition: all stocks are australian stocks

wb1 = openpyxl.load_workbook('test.xlsx')

sheet1 = wb1.get_sheet_by_name('Sheet1')

stocksString = ''

for x in range (2,21):
    #print(sheet1.cell(row=1,column=x).value)
    if (x != 20):
        stocksString = stocksString + sheet1.cell(row=1,column=x).value +'.AX+' 
    else: 
        stocksString = stocksString + sheet1.cell(row=1,column=x).value +'.AX' 

#print (stocksString)

#grab csv with yahoo finance API

http = urllib3.PoolManager()


url = 'http://finance.yahoo.com/d/quotes.csv?s=' + stocksString + '&f=sl1'

response = http.request('GET', url)

with open('quotes.csv', 'wb') as f:
    f.write(response.data)

response.release_conn()

#print (response.data)

df1=pd.read_csv("quotes.csv",header=None)

print(df1)

i = 0

while (i < 3):
    print(df1.ix[i,1])
    i=i+1

wb = openpyxl.load_workbook('invest.xlsx')

#file = 'invest.xlsx'
#print (wb)
sheet = wb.get_sheet_by_name('Sheet1')

#c = sheet['A1']

#print (c.value)

#print ("hello world")

old_sheet = wb.get_sheet_by_name('Sheet1')
old_sheet.title = 'Sheet1.5'
max_row = old_sheet.max_row
max_col = old_sheet.max_column
wb.create_sheet('Sheet1new', 0)

print ("the max column is:", max_col)

print ("the max row is:", max_row)

new_sheet = wb.get_sheet_by_name('Sheet1new')

# Do the header.
for col_num in range(1, max_col+1):
    new_sheet.cell(row=1, column=col_num).value = old_sheet.cell(row=1, column=col_num).value

# The row to be inserted. We're manually populating each cell.

# Now do the rest of it. Note the row offset.
for row_num in range(2, max_row+1):
   for col_num in range (1, max_col+1):
        new_sheet.cell(row = (row_num + 1), column = col_num).value = old_sheet.cell(row = row_num, column = col_num).value

dateToday = time.strftime("%d/%m/%Y")

i = 2
j = 0

new_sheet.cell(row=2, column=1).value = dateToday 

while (i < 21):
    new_sheet.cell(row=2, column=i).value = df1.ix[j,1]
    i = i+1
    j = j+1


#new_sheet.cell(row=2, column=2).value = df1.ix[0,1]

wb.remove_sheet(old_sheet)
new_sheet.title = 'Sheet1'

wb.save('invest.xlsx')

